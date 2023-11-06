#-----------------------------------------------------
# Intel Corporation - SMG SS LATAM E&G
# Web Scrapping from Colombia Compra Eficiente POs
#
# Author: Kimberly Orozco-Retana ITS.Intern
#-----------------------------------------------------

#import libraries
from bs4 import BeautifulSoup #for scrapping
import pandas as pd #for data management
import requests #for urls
import time
from openpyxl import Workbook, load_workbook


path = path = r"C:\Users\kdorozco\OneDrive - Intel Corporation\Desktop\Automatization_Projects\Items_ETP3.xlsx"


def PO_details(items_PO):
    data = pd.read_excel(items_PO, sheet_name='Órdenes')
    orders = data.Orden
    base_url = 'https://www.colombiacompra.gov.co/tienda-virtual-del-estado-colombiano/ordenes-compra/'
    PO_info = {}
    d = 1

    book = load_workbook(items_PO)
    sheet = book['Órdenes']
    print(sheet)

    k = 2
    for i in orders:
        per = str(round(100*d/len(orders),2))+'%' #shows % progress in scrapping
        PO_info[i] = {}
        url = base_url + str(i)
        response = requests.get(url)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            entidad = str(soup.find('label', string='Entidad').find_next('span', class_='oc-span').text.strip())
            justificacion = str(soup.find('label', string='Justificación').find_next('span', class_='oc-span').text.strip())
            nombre = str(soup.find('label', string='Nombre').find_next('span', class_='oc-span').text.strip())
            fecha = str(soup.find('label', string='Fecha de la orden').find_next('span', class_='oc-span').text.strip())

            sheet.cell(k, 2).value = entidad
            sheet.cell(k, 3).value = fecha
            sheet.cell(k, 5).value = nombre
            sheet.cell(k, 9).value = justificacion
            k += 1
            print(per)
            d += 1
            time.sleep(2)

        else:
            print('Failed to retrieve the webpage. Status code: ', response.status_code)
            print('Please increase the delay between requests.')

            
    
    book.save(items_PO)

PO_details(path)